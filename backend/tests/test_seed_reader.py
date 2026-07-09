from __future__ import annotations

from pathlib import Path

import openpyxl

from app.seed.parse import RowKind
from app.seed.reader import detect_building_type, read_workbook


def test_detect_building_type() -> None:
    assert detect_building_type(r"c:\x\...жилые здания....xlsx") == "residential"
    assert detect_building_type("офисные, ТРЦ.xlsx") == "office"
    assert detect_building_type("социальные объекты.xlsx") == "social"


def _make_book(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Производители"
    ws["A5"] = "№"
    ws["B5"] = "Наименование"
    ws["C5"] = "Комфорт"
    ws.cell(row=5, column=100, value="ФАНТОМ")  # далёкая колонка — не должна читаться
    ws["A6"] = "1. Оборудование"               # heading (B пусто)
    ws["A7"] = 1
    ws["B7"] = "Насос"
    ws["C7"] = "\n Grundfos"                    # ведущий \n снимается
    ws["A8"] = 2
    ws["B8"] = "Клапан"
    # вертикальный merge C7:C8 — значение из C7 должно попасть в позицию строки 8
    ws.merge_cells("C7:C8")
    hidden = wb.create_sheet("Vendor list")
    hidden.sheet_state = "hidden"
    wb.save(path)


def test_read_workbook(tmp_path: Path) -> None:
    p = tmp_path / "тест жилые.xlsx"
    _make_book(p)
    data = read_workbook(str(p))

    assert data.building_type == "residential"
    assert data.hidden_sheets == ["Vendor list"]
    # границы колонок по шапке, НЕ по max_column (фантом в C=col100 игнорируется)
    assert [c.col for c in data.classes] == [3]
    assert data.classes[0].segment_name == "Комфорт"

    positions = [r for r in data.rows if r.row.kind is RowKind.POSITION]
    assert len(positions) == 2
    assert positions[0].cells[3] == "Grundfos"          # ведущий \n снят
    assert positions[1].cells[3] == "Grundfos"          # merge C7:C8 развёрнут
