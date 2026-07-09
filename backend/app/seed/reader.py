"""Чтение сид-книг openpyxl: видимый лист, детекция шапки, границы колонок,
разворот объединённых ячеек для колонок классов. Классификация строк — из parse."""

from __future__ import annotations

import os
from dataclasses import dataclass

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from .parse import RowClass, RowKind, SeedError, _text, classify_row

_BT_BY_SUBSTR = (("жил", "residential"), ("офис", "office"), ("социальн", "social"))
_MAX_HEADER_SCAN = 30
_MAX_CLASS_COL = 40  # верхняя граница поиска классов (защита от фантомных 16384)


@dataclass(frozen=True)
class ClassColumn:
    col: int
    segment_name: str
    group_name: str | None


@dataclass
class ReadRow:
    row_no: int
    row: RowClass
    cells: dict[int, str]


@dataclass
class SheetData:
    building_type: str
    sheet_title: str
    hidden_sheets: list[str]
    classes: list[ClassColumn]
    rows: list[ReadRow]


def detect_building_type(path: str) -> str:
    low = os.path.basename(path).lower()
    for sub, code in _BT_BY_SUBSTR:
        if sub in low:
            return code
    raise SeedError(f"Не определён тип объекта по имени файла: {path!r}")


def _merge_anchors(ws: Worksheet, max_col: int) -> dict[tuple[int, int], tuple[int, int]]:
    """(row,col)->(anchor_row,anchor_col) только для merge в колонках классов (col>=3)."""
    lut: dict[tuple[int, int], tuple[int, int]] = {}
    for mr in ws.merged_cells.ranges:
        if mr.min_col < 3 or mr.min_col > max_col:
            continue  # пропускаем заголовочные A:H и фантомные далёкие merge
        anchor = (mr.min_row, mr.min_col)
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                lut[(r, c)] = anchor
    return lut


def _resolve(ws: Worksheet, lut: dict[tuple[int, int], tuple[int, int]], r: int, c: int) -> object:
    anchor = lut.get((r, c))
    if anchor is not None:
        return ws.cell(row=anchor[0], column=anchor[1]).value
    return ws.cell(row=r, column=c).value


def _nearest_left(ws: Worksheet, row: int, col: int) -> str | None:
    for c in range(col, 2, -1):
        v = _text(ws.cell(row=row, column=c).value)
        if v:
            return v
    return None


def read_workbook(path: str) -> SheetData:
    bt = detect_building_type(path)
    wb = openpyxl.load_workbook(path, data_only=True)
    hidden = [ws.title for ws in wb.worksheets if ws.sheet_state != "visible"]
    visibles = [ws for ws in wb.worksheets if ws.sheet_state == "visible"]
    if len(visibles) != 1:
        raise SeedError(f"{path}: ожидался один видимый лист, найдено {len(visibles)}")
    ws = visibles[0]

    anchor_row = next(
        (r for r in range(1, min(ws.max_row, _MAX_HEADER_SCAN) + 1)
         if _text(ws.cell(row=r, column=1).value) == "№"),
        None,
    )
    if anchor_row is None:
        raise SeedError(f"{path}: не найдена строка-якорь шапки (№ в колонке A)")

    below_a = _text(ws.cell(row=anchor_row + 1, column=1).value)
    below_b = _text(ws.cell(row=anchor_row + 1, column=2).value)
    below_classes = any(
        _text(ws.cell(row=anchor_row + 1, column=c).value) for c in range(3, _MAX_CLASS_COL)
    )
    two_level = below_a == "" and below_b == "" and below_classes
    class_row = anchor_row + 1 if two_level else anchor_row
    group_row = anchor_row if two_level else None
    data_start = anchor_row + 2 if two_level else anchor_row + 1

    classes: list[ClassColumn] = []
    c = 3
    while c < _MAX_CLASS_COL:
        label = _text(ws.cell(row=class_row, column=c).value)
        if label == "":
            break
        group = _nearest_left(ws, group_row, c) if group_row is not None else None
        classes.append(ClassColumn(c, label, group))
        c += 1
    if not classes:
        raise SeedError(f"{path}: не найдено ни одной колонки класса на строке {class_row}")

    last_col = classes[-1].col
    lut = _merge_anchors(ws, last_col)

    rows: list[ReadRow] = []
    for r in range(data_start, ws.max_row + 1):
        rc = classify_row(ws.cell(row=r, column=1).value, ws.cell(row=r, column=2).value, row_no=r)
        cells: dict[int, str] = {}
        if rc.kind is RowKind.POSITION:
            for cc in classes:
                s = _text(_resolve(ws, lut, r, cc.col))
                if s:
                    cells[cc.col] = s
        rows.append(ReadRow(r, rc, cells))
    return SheetData(bt, ws.title, hidden, classes, rows)
