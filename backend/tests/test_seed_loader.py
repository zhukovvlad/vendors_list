from __future__ import annotations

from pathlib import Path

import openpyxl

from app.seed.loader import build_load


def _make(path: Path, title: str, sheetname: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheetname
    ws["A5"] = "№"
    ws["B5"] = "Наименование"
    ws["C5"] = "Комфорт"
    ws["A6"] = "1. Оборудование"
    ws["A7"] = "1.1. Отопление"
    ws["A8"] = 1
    ws["B8"] = "Насос"
    ws["C8"] = "Ридан, ТеплоСила*"
    ws["A9"] = 2
    ws["B9"] = "Клапан"
    ws["C9"] = "-"
    wb.save(path)


def test_build_load_counts_and_stars(tmp_path: Path) -> None:
    p = tmp_path / "тест жилые.xlsx"
    _make(p, "жилой", "Производители")
    plan = build_load([str(p)])

    assert len(plan.positions) == 2
    # listing: 2 вендора (Насос) + 1 мета '-' (Клапан) = 3
    assert len(plan.listings) == 3
    assert plan.vendors == {"Ридан": False, "ТеплоСила": True}
    # дерево: 1 и 1.1
    assert [n.number for n in plan.categories] == [(1,), (1, 1)]
    dash = [ln for ln in plan.listings if ln.status == "not_applicable"]
    assert len(dash) == 1 and dash[0].vendor_name is None


def test_build_load_is_db_free(tmp_path: Path) -> None:
    # build_load не должен требовать DATABASE_URL — вызов не падает без БД
    p = tmp_path / "тест социальные.xlsx"
    _make(p, "соц", "Оборудование ")
    plan = build_load([str(p)])
    assert plan.report.files[0].building_type == "social"


def _make_dup(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Производители"
    ws["A5"] = "№"
    ws["B5"] = "Наименование"
    ws["C5"] = "Комфорт"
    ws["A6"] = "1. Оборудование"
    ws["A7"] = 1
    ws["B7"] = "Насос"
    ws["C7"] = "Ридан, Ридан"  # повтор бренда в одной ячейке
    wb.save(path)


def test_build_load_dedupes_repeated_vendor_in_cell(tmp_path: Path) -> None:
    # uq_listing_cell_vendor: один vendor_id дважды в ячейке недопустим — дубль
    # пропускается, остаётся один listing + предупреждение
    p = tmp_path / "тест жилые.xlsx"
    _make_dup(p)
    plan = build_load([str(p)])
    allowed = [ln for ln in plan.listings if ln.status == "allowed"]
    assert len(allowed) == 1
    assert plan.vendors == {"Ридан": False}
    assert any("повтор вендора" in w for f in plan.report.files for w in f.warnings)


def test_build_load_sort_order_resets_per_file(tmp_path: Path) -> None:
    # §7: sort_order позиции — порядок В ПРЕДЕЛАХ ФАЙЛА, не глобальный индекс.
    p1 = tmp_path / "тест жилые.xlsx"
    p2 = tmp_path / "тест социальные.xlsx"
    _make(p1, "жилой", "Производители")
    _make(p2, "соц", "Производители")
    plan = build_load([str(p1), str(p2)])

    klapan_positions = [p for p in plan.positions if p.name == "Клапан"]
    # обе позиции 'Клапан' (файл1 и файл2) должны иметь sort_order=1 (второй в своём файле)
    assert len(klapan_positions) == 2
    assert [p.sort_order for p in klapan_positions] == [1, 1]

    насос_positions = [p for p in plan.positions if p.name == "Насос"]
    assert len(насос_positions) == 2
    assert [p.sort_order for p in насос_positions] == [0, 0]

    # global pos_key продолжает расти сквозным счётчиком (это НЕ sort_order)
    assert [p.pos_key for p in plan.positions] == [1, 2, 3, 4]
